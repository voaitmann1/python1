import pandas as pd
file = 'D:\\temp\\example1.xls'
xl = pd.ExcelFile(file)
df1 = xl.parse('Sheet1')
print(df1)
print(df1.columns[1])
print(df1[1][1])
#writer = pd.ExcelWriter(file, engine='xlsxwriter')
#writer = pd.ExcelWriter(file, engine='xlswriter')
#df1.to_excel(writer, 'Sheet2')
#
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
# Load in the workbook
wb = load_workbook(file)
# Get sheet names
print(wb.get_sheet_names())
#
for i in range(1, 4):
     print(i, sheet.cell(row=i, column=2).value)
